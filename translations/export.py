#!/usr/bin/python3

import os
import sys
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.comments import Comment

def print_help():
    help_message = """
Usage:
    python export.py [INPUT_DIRECTORY] [OUTPUT_XLSX_FILE]

Description:
    This script extracts string resources from Android strings.xml files
    and exports them to an Excel file. It supports plurals and includes formatting and hiding of non-translatable rows.

    - INPUT_DIRECTORY: Path to the root directory containing the strings.xml files (in 'values-{LANGUAGE}' subfolders)
      (default: './app/src/main/res').
    - OUTPUT_XLSX_FILE: Path to the output Excel file (default: 'translations.xlsx').

Features:
    - Extracts all string resources into a single Excel file with columns for each language.
    - Hides rows for non-translatable strings.
    - Applies formatting to the Excel file:
        * Highlighting of the header row and empty cells.
        * "Key" and "Translatable" columns are hidden.
    - Sorted by the "Key" column alphabetically.

Example:
    python export.py ./app/src/main/res translations.xlsx

Help:
    Use the -h or --help flag to display this help message.
"""
    print(help_message)

def process_plurals_element(element, folderName, string_map, row_hidden_map):
    """
    Process the <plurals> element in the XML file and ensure every plural has
    rows for all quantities (few, many, one, other, two, zero).
    """
    key = element.attrib['name']
    quantities = ["few", "many", "one", "other", "two", "zero"]  # All possible plural quantities
    # Ensure every quantity is included
    for quantity in quantities:
        plural_key = f"{key}_PLURALS_{quantity}"

        # Check if the quantity exists in the XML; if not, set to None
        text = None
        for item in element.findall("item"):
            if item.attrib.get("quantity") == quantity:
                text = item.text
                break

        # Add the quantity to the string_map under the appropriate language folder
        if plural_key not in string_map:
            string_map[plural_key] = {}

        string_map[plural_key][folderName] = text
        row_hidden_map[plural_key] = element.attrib.get('translatable', 'true') == 'false'


def get_all_xml_files(directory):
    """
    Retrieve all strings.xml files from the given directory.
    """
    all_files = []  # List to store all XML files
    for root, _, files in os.walk(directory):
        for filename in files:
            if filename == "strings.xml":
                filepath = os.path.join(root, filename)
                all_files.append(filepath)
    return all_files

def process_xml_files(all_files, string_map, row_hidden_map, folderNames):
    """
    Process each XML file to populate string_map and row_hidden_map.
    """
    total_files = len(all_files)
    print(f"Found {total_files} files to process.")

    for idx, filepath in enumerate(all_files, start=1):
        folderName = os.path.basename(os.path.dirname(filepath))
        folderNames.append(folderName)

        # Print progress
        print(f"Processing file {idx}/{total_files}: {filepath}")

        tree = ET.parse(filepath, parser=ET.XMLParser(encoding="UTF-8"))
        root = tree.getroot()

        for element in root:
            if element.tag == "string":
                key = element.attrib['name']
                text = element.text
                hidden = element.attrib.get('translatable', 'true') == 'false'

                dict = string_map.get(key, {})
                dict[folderName] = text
                string_map[key] = dict
                row_hidden_map[key] = hidden
            elif element.tag == "plurals":
                process_plurals_element(element, folderName, string_map, row_hidden_map)

def create_dataframe(string_map, row_hidden_map, folderNames):
    """
    Convert the string_map into a pandas DataFrame.
    """
    header = ["Key", "Translatable"] + sorted(folderNames)
    df = pd.DataFrame(columns=header)
    rows = []

    for key, value in string_map.items():
        value["Key"] = key
        value["Translatable"] = "false" if row_hidden_map.get(key, False) else "true"
        rows.append(value)

    df = pd.concat([df, pd.DataFrame(rows)], ignore_index=True)
    df.sort_values(by="Key", inplace=True)
    return df

def format_excel(output_file, df, row_hidden_map):
    """
    Apply formatting to the Excel file and save it.
    """
    wb = load_workbook(output_file)
    ws = wb.active

    # Apply formatting (borders, header styles, etc.)
    ws.freeze_panes = ws['D2']
    fixed_width = 50
    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = fixed_width

    # Hide Translatable column
    ws.column_dimensions['B'].hidden = True

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(name="Calibri", bold=True)
    cell_font = Font(name="Calibri")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Handle empty cells highlighting
    empty_cell_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    empty_cell_plural_fill = PatternFill(start_color="ffffcc", end_color="ffffcc", fill_type="solid")
    text_style = NamedStyle(name="text_style", number_format="@")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.style = text_style
            cell.font = cell_font
            if cell.value is None or str(cell.value).strip() == "":
                if "_PLURALS_" not in row[0].value:
                    cell.fill = empty_cell_fill
                elif row[2].value is None:
                    comment_text = f"There is no translation needed in the default language for the quantity '{row[0].value.split('_PLURALS_')[1]}'.\nIf this language needs a different translation for this quantity add it, otherwise ignore this cell."
                    cell.comment = Comment(comment_text, "Plurals Checker")
                    cell.fill = empty_cell_plural_fill

    # Check and modify missing plural values in the default language column (3rd column)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        key = row[0].value
        if "_PLURALS_" in key:
            default_language_cell = row[2]  # Get the cell in the default language column (3rd column)
            if default_language_cell.value is None or str(default_language_cell.value).strip() == "" or str(default_language_cell.value).startswith("Key:"):
                comment_text = f"There is no translation needed in the default language for the quantity '{key.split('_PLURALS_')[1]}'"
                default_language_cell.comment = Comment(comment_text, "Plurals Checker")
                default_language_cell.fill = empty_cell_plural_fill

    # Thin borders for all cells
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            cell.border = thin_border

    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for idx, key in enumerate(df["Key"], start=2):
        if row_hidden_map.get(key, False):
            ws.row_dimensions[idx].hidden = True
            # Apply gray highlight to all cells in the row
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=idx, column=col)
                cell.fill = gray_fill
                cell.comment = Comment("This text is not translatable, ignore it", "Plurals Checker")

    # Apply an AutoFilter to the third column (default language column)
    start_column = 'A'
    end_column = ws.max_column
    ws.auto_filter.ref = f"{start_column}1:{ws.cell(row=1, column=end_column).coordinate}"

    wb.save(output_file)

def main():
    if len(sys.argv) > 1 and sys.argv[1] in ["-h", "--help"]:
        print_help()
        sys.exit(0)

    default_directory = './app/src/main/res'
    directory = sys.argv[1] if len(sys.argv) > 1 else default_directory

    default_output_file = "translations.xlsx"
    output_file = sys.argv[2] if len(sys.argv) > 2 else default_output_file

    string_map = {}
    row_hidden_map = {}
    folderNames = []

    all_files = get_all_xml_files(directory)

    # Process XML files
    process_xml_files(all_files, string_map, row_hidden_map, folderNames)

    print("Processing complete.")

    # Create dataframe and generate Excel
    print("Generating Excel file...")
    df = create_dataframe(string_map, row_hidden_map, folderNames)
    df.to_excel(output_file, index=False)

    # Apply formatting and save
    format_excel(output_file, df, row_hidden_map)

    print(f"Export completed. Check {output_file} for the output.")

if __name__ == "__main__":
    main()
